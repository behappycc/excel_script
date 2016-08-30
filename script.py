# Python module
import json

# OpenPyxXL module
import openpyxl

def read_config():
    with open('config.json', 'r') as file:
        config = json.load(file)
    return config

def main():
    config = read_config()
    print(config['INVMB'])
    wb_INVMB = openpyxl.load_workbook(config['INVMB'])
    print ('open: {}'.format(config['INVMB']))

    wb_INVMD = openpyxl.load_workbook(config['INVMD'])
    print ('open: {}'.format(config['INVMD']))

    '''
    wb_INVMB = openpyxl.load_workbook('INVMB.xlsx')
    print ('open: {}'.format('INVMB.xlsx'))

    wb_INVMD = openpyxl.load_workbook('INVMD.xlsx')
    print ('open: {}'.format('INVMD.xlsx'))
    '''


    ws_INVMB = wb_INVMB.get_sheet_by_name('INVMB')
    ws_INVMD = wb_INVMD.get_sheet_by_name('Sheet1')

    INVMD_MD001 = {}
    INVMD_MD001_index = []
    for row in ws_INVMD.iter_rows('N2:N136'):
        for cell in row:
            INVMD_MD001_index.append(cell.value)
            INVMD_MD001[cell.value] = cell.row
    
    for row in ws_INVMB.iter_rows('N2:N551'):
        for cell in row:
            # cell attributes: row, value
            
            # INVMB exculde RA
            if not (cell.value).startswith('RA'):
                if cell.value in INVMD_MD001_index:
                    index = INVMD_MD001[cell.value]
                    ws_INVMB['FL' + str(cell.row)].value = ws_INVMD['O' + str(index)].value 
                else:
                    ws_INVMB['FL' + str(cell.row)].value = ws_INVMB['Q' + str(cell.row)].value

            # INVMB RA
            else:
                if cell.value in INVMD_MD001_index:
                    index = INVMD_MD001[cell.value]
                    ws_INVMB['FL' + str(cell.row)].value = ws_INVMB['Q' + str(cell.row)].value            
                else:
                    ws_INVMB['FL' + str(cell.row)].value = ws_INVMD['O' + str(index)].value

    print ('create new INVMB: {}'.format('new_INVMD.xlsx'))
    wb_INVMB.save('new_INVMB.xlsx')

if __name__ == '__main__':
    main()